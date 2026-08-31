#!/usr/bin/env python3
"""Convert markdown content to XLSX using openpyxl."""
import sys
import argparse
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def markdown_to_xlsx(md_path: str, output_path: str) -> bool:
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        sys.stderr.write(f"读取 markdown 失败: {e}\n")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Content"

    # Try to detect tables (lines with | separators)
    lines = content.split("\n")
    table_rows = []
    in_table = False

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            if not in_table:
                in_table = True
                table_rows = []
            # Parse table row
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            table_rows.append(cells)
        else:
            if in_table and table_rows:
                # Flush table
                _write_table(ws, table_rows)
                ws.move_row(len(table_rows) + 1)
                table_rows = []
                in_table = False
            # Non-table content as single cell
            if stripped:
                ws.append([stripped])

    if in_table and table_rows:
        _write_table(ws, table_rows)

    return True


def _write_table(ws, rows):
    """Write a table row-by-row with styling."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    cell_font = Font(size=10)

    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            c = ws.cell(row=i + 1, column=j + 1, value=cell)
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if i == 0:
                c.font = header_font
            else:
                c.font = cell_font
        # Auto-adjust column width
        for j, cell in enumerate(row):
            col_letter = ws.cell(row=1, column=j + 1).column_letter
            max_len = min(len(cell) + 2, 50)
            ws.column_dimensions[col_letter].width = max_len


def main():
    parser = argparse.ArgumentParser(description="Markdown to XLSX converter")
    parser.add_argument("input", help="Input markdown file path")
    parser.add_argument("output", help="Output XLSX file path")
    args = parser.parse_args()

    success = markdown_to_xlsx(args.input, args.output)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
