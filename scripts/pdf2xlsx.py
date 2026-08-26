#!/usr/bin/env python3
"""PDF 转 XLSX 工具 - 使用 pdfminer + openpyxl"""
import sys
import argparse
from pdfminer.high_level import extract_text
from openpyxl import Workbook


def pdf_to_xlsx(pdf_path: str, output_path: str) -> bool:
    text = extract_text(pdf_path)
    if not text.strip():
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Content"

    for line in text.splitlines():
        # 尝试按空白字符分割成列
        cells = line.split()
        for col_idx, cell_text in enumerate(cells, 1):
            ws.cell(row=ws.max_row, column=col_idx, value=cell_text)
        ws.row_dimensions[ws.max_row].height = 18

    wb.save(output_path)
    return True


def main():
    parser = argparse.ArgumentParser(description="PDF to XLSX converter")
    parser.add_argument("input", help="Input PDF file path")
    parser.add_argument("output", help="Output XLSX file path")
    args = parser.parse_args()

    success = pdf_to_xlsx(args.input, args.output)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
